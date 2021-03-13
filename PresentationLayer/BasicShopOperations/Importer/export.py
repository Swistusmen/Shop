from openpyxl import Workbook
import os

from DataAccessLayer import crud
from DataAccessLayer import database

def get_db():
    db=database.SessionLocal()
    try:
        yield db
    finally:
        db.close()

def export_users(db):
    dirs=os.listdir()
    if "users.xlxs" in dirs:
        os.remove("users.xlsx")
    wb= Workbook()
    ws= wb.active
    ws["A1"]="id"
    ws["B1"]="name" 
    ws["C1"]="surname"
    ws["D1"]="email" 
    ws["E1"]="wallet" 
    ws["F1"]="password" 
    ws["G1"]="is_admin"    
    list_of_users=crud.get_users(db)
    current_row=2
    for i in list_of_users:
        ws.cell(row=current_row, column=1, value=i.id)
        ws.cell(row=current_row, column=2, value=i.name)
        ws.cell(row=current_row, column=3, value=i.surname)
        ws.cell(row=current_row, column=4, value=i.email)
        ws.cell(row=current_row, column=5, value=i.wallet)
        ws.cell(row=current_row, column=6, value=i.password)
        ws.cell(row=current_row, column=7, value=i.is_admin)
        current_row+=1
    wb.save("users.xlsx")

def export_products(db):
    dirs=os.listdir()
    if "products.xlxs" in dirs:
        os.remove("products.xlsx")
    wb= Workbook()
    ws= wb.active
    ws["A1"]="id"
    ws["B1"]="name" 
    ws["C1"]="description" 
    ws["D1"]="number" 
    ws["E1"]="price" 
    ws["F1"]="user_id"     
    list_of_users=crud.get_all_products(db)
    current_row=2
    for i in list_of_users:
        ws.cell(row=current_row, column=1, value=i.id)
        ws.cell(row=current_row, column=2, value=i.name)
        ws.cell(row=current_row, column=3, value=i.description)
        ws.cell(row=current_row, column=4, value=i.number)
        ws.cell(row=current_row, column=5, value=i.price)
        ws.cell(row=current_row, column=6, value=i.user_id)
        current_row+=1
    wb.save("products.xlsx")

def export_orders(db):
    dirs=os.listdir()
    if "orders.xlsx" in dirs:
        os.remove("orders.xlsx")
    wb= Workbook()
    ws= wb.active
    ws["A1"]="id"
    ws["B1"]="isFinished" 
    ws["C1"]="items" 
    ws["D1"]="user_id"    
    list_of_users=crud.get_orders(db)
    current_row=2
    for i in list_of_users:
        ws.cell(row=current_row, column=1, value=i.id)
        ws.cell(row=current_row, column=2, value=i.isFinished)
        ws.cell(row=current_row, column=3, value=i.items)
        ws.cell(row=current_row, column=4, value=i.user_id)
        current_row+=1
    wb.save("orders.xlsx")

   

