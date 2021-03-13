from openpyxl import Workbook, load_workbook
import os

from DataAccessLayer import crud
from DataAccessLayer import database
from DataAccessLayer import  models, schemas

def get_db():
    db=database.SessionLocal()
    try:
        yield db
    finally:
        db.close()

def start_db_with_users(db):
    wb= load_workbook("users.xlsx")
    ws= wb.active

    my_users=[]
    current_row=2
    while ws.cell(row=current_row, column=1).value!=None:
        my_users.append(schemas.UserCreate(
            id=ws.cell(row=current_row, column=1).value,
            name=ws.cell(row=current_row, column=2).value,
            surname=ws.cell(row=current_row, column=3).value,
            email=ws.cell(row=current_row, column=4).value,
            wallet=ws.cell(row=current_row, column=5).value,
            password=ws.cell(row=current_row, column=6).value,
            is_admin=ws.cell(row=current_row, column=7).value,
            is_disabled=ws.cell(row=current_row, column=8).value,
            ))
        current_row+=1

    print(len(my_users))
    for i in my_users:
        crud.create_user(db=db,user=i)
    db.close()
    return

def start_db_with_products(db):
    wb= load_workbook("products.xlsx")
    ws= wb.active

    my_products=[]
    current_row=2
    while ws.cell(row=current_row, column=1).value!=None:
        my_products.append(schemas.Product(
            id=ws.cell(row=current_row, column=1).value,
            name=ws.cell(row=current_row, column=2).value,
            description=ws.cell(row=current_row, column=3).value,
            number=ws.cell(row=current_row, column=4).value,
            price=ws.cell(row=current_row, column=5).value,
            user_id=ws.cell(row=current_row, column=6).value,
            ))
        current_row+=1

    for i in my_products:
        crud.create_product(db=db,product=i,shop_id=1)
    db.close()
    return  

def start_db_with_orders(db):
    wb= load_workbook("orders.xlsx")
    ws= wb.active

    my_orders=[]
    current_row=2
    while ws.cell(row=current_row, column=1).value!=None:
        my_orders.append(schemas.Order(
            id=ws.cell(row=current_row, column=1).value,
            isFinished=ws.cell(row=current_row, column=2).value,
            products=ws.cell(row=current_row, column=3).value,
            owner_id=ws.cell(row=current_row, column=4).value,
            ))
        current_row+=1

    print(len(my_orders))
    for i in my_orders:
        crud.create_order(db=db,order=i)
    db.close()
    return
